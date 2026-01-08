import pandas as pd
import numpy as np
from fuzzywuzzy import fuzz
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from tqdm import tqdm
import re
from datetime import datetime, timedelta
from functools import lru_cache  
import time
import os


@lru_cache(maxsize=None)
def standardize_date(date_str):
    if pd.isna(date_str):
        return None
    if isinstance(date_str, str):
        # Try parsing string dates
        for fmt in ('%Y-%m-%d %H:%M:%S', '%m/%d/%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
    elif isinstance(date_str, pd.Timestamp):
        # If it's already a Timestamp, just return the date part
        return date_str.date()
    return None

@lru_cache(maxsize=None)
def preprocess_name(name):
    # Remove titles, suffixes, etc.
    name = re.sub(r'\b(Mr|Mrs|Ms|Dr|Jr|Sr|I{1,3})\b\.?\s*', '', str(name))
    # Convert to lowercase and remove extra spaces
    return ' '.join(name.lower().split())

def compare_names(name1, name2):
    # Preprocess names
    name1 = preprocess_name(name1)
    name2 = preprocess_name(name2)
    
    # Check for exact match after preprocessing
    if name1 == name2:
        return 100
    
    # Split into tokens
    tokens1 = name1.split()
    tokens2 = name2.split()
    
    # Compare last names (assuming last token is last name)
    last_name_match = fuzz.ratio(tokens1[-1], tokens2[-1])
    
    # Compare first names (assuming first token is first name)
    first_name_match = fuzz.ratio(tokens1[0], tokens2[0])
    
    # Handle middle names/initials
    middle_match = 0
    if len(tokens1) > 2 and len(tokens2) > 2:
        # If both have middle names/initials
        middle_match = fuzz.ratio(' '.join(tokens1[1:-1]), ' '.join(tokens2[1:-1]))
    elif len(tokens1) > 2:
        # If only name1 has a middle name/initial, compare it with the first letter of name2's first name
        middle_match = fuzz.ratio(tokens1[1][0], tokens2[0][0])
    elif len(tokens2) > 2:
        # If only name2 has a middle name/initial, compare it with the first letter of name1's first name
        middle_match = fuzz.ratio(tokens1[0][0], tokens2[1][0])
    
    # Compare full names
    full_name_match = fuzz.token_sort_ratio(name1, name2)
    
    # Weighted score optimized for name matching:
    # 40% last_name (highest signal) + 30% first_name + 10% middle_initial + 20% token_sort (adjust as needed)
    score = (0.4 * last_name_match) + (0.3 * first_name_match) + (0.1 * middle_match) + (0.2 * full_name_match)
    
    return score

def compare_spreadsheets(file1, file2, overall_threshold=69, weights=(0.6, 0.3, 0.1)):
    """Compares two spreadsheets using vectorized operations for faster processing."""
    # Production scale: loops through 10k source records, each compared against 150k target
    # Vectorized exact match first → fuzzy only when needed

    try:
        df1 = pd.read_csv(file1, encoding='latin1')
        df2 = pd.read_csv(file2, encoding='latin1')
    except Exception as e:
        print(f"Error reading CSV files: {e}")
        return None

    match_columns = ['Individual Name', 'Individual Birthdate', 'Individual Gender']

    if not all(col in df1.columns for col in match_columns) or not all(col in df2.columns for col in match_columns):
        print("Error: Not all specified columns exist in both spreadsheets.")
        return None
    
    # Preprocessing columns used for matching
    for df in [df1, df2]:
        df['preprocessed_name'] = df['Individual Name'].apply(preprocess_name)
        df['standardized_date'] = df['Individual Birthdate'].apply(standardize_date)
        df['gender_lower'] = df['Individual Gender'].str.lower().str.strip()

    def vectorized_fuzzy_match(names1, names2):
        return np.array([compare_names(n1, n2) for n1, n2 in zip(names1, names2)])

    matches = []
    for _, row1 in tqdm(df1.iterrows(), total=len(df1), desc="Matching rows"):
        # Vectorized exact match
        exact_matches = (df2['preprocessed_name'] == row1['preprocessed_name']) & \
                        (df2['standardized_date'] == row1['standardized_date']) & \
                        (df2['gender_lower'] == row1['gender_lower'])
        
        if exact_matches.any():
            best_match = df2[exact_matches].iloc[0]
            best_score = 100
        else:
            # Vectorized fuzzy matching
            name_scores = vectorized_fuzzy_match(np.full(len(df2), row1['preprocessed_name']), df2['preprocessed_name'])
            date_scores = (df2['standardized_date'] == row1['standardized_date']).astype(int) * 100
            gender_scores = (df2['gender_lower'] == row1['gender_lower']).astype(int) * 100
            
            overall_scores = weights[0] * name_scores + weights[1] * date_scores + weights[2] * gender_scores
            
            if overall_scores.max() >= overall_threshold:
                best_match_idx = overall_scores.argmax()
                best_match = df2.iloc[best_match_idx]
                best_score = overall_scores[best_match_idx]
            else:
                continue

        match_result = {f"File1_{col}": row1[col] for col in match_columns}
        match_result["File1_Individual: ID Number"] = row1["Individual Name: ID Number"]
        match_result.update({f"File2_{col}": best_match[col] for col in df2.columns})
        match_result['match_score'] = best_score
        matches.append(match_result)
        # Note: For varied datasets, use row1.get("Individual: ID Number", row1.get("ID Number", "")) to handle missing columns gracefully

    return pd.DataFrame(matches)

def export_to_excel(df, output_file, highlight_column):
    if df is None or df.empty:
        print("No matches found. No output file will be created.")
        return

    # Add SourceFile_/TargetFile_ prefixes
    df.rename(columns=lambda x: f"SourceFile_{x[6:]}" if x.startswith("File1_") else x, inplace=True)
    df.rename(columns=lambda x: f"TargetFile_{x[6:]}" if x.startswith("File2_") else x, inplace=True)

    # Move match_score to left, sort fuzzy first
    cols = df.columns.tolist()
    cols.insert(0, cols.pop(cols.index('match_score')))
    df = df.reindex(columns=cols)
    
    df['is_fuzzy'] = df['match_score'] < 100
    df_sorted = df.sort_values(['is_fuzzy', 'match_score'], ascending=[False, False]).drop('is_fuzzy', axis=1)
    
    # Save + style Excel
    df_sorted.to_excel(output_file, sheet_name='Matches', index=False, engine='openpyxl')
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Define styles
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    fuzzy_fill = PatternFill(start_color='FFD580', end_color='FFD580', fill_type='solid')
    id_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    header_font = Font(bold=True)

    # Headers + column widths, Fuzzy matches (score<100) sorted FIRST - orange highlight; ID Number columns yellow highlighted; Frozen headers; auto-sized columns
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = (max_length + 2) * 1.2

    ws.freeze_panes = 'A2'

    # ID columns + formatting
    source_id_cols = [col for col in df_sorted.columns if col.startswith('SourceFile_') and 'ID Number' in col]
    target_id_cols = [col for col in df_sorted.columns if col.startswith('TargetFile_') and 'ID Number' in col]

    for row in range(2, ws.max_row + 1):
        match_score = ws[f'A{row}'].value
        if match_score < 100:
            for col in range(2, ws.max_column + 1):
                if str(ws.cell(row=1, column=col).value).startswith('SourceFile_'):
                    ws.cell(row=row, column=col).fill = fuzzy_fill
        
        for col_name in source_id_cols + target_id_cols:
            col_letter = get_column_letter(df_sorted.columns.get_loc(col_name) + 1)
            ws[f'{col_letter}{row}'].fill = id_fill

    wb.save(output_file)
    print(f"Results exported to {output_file} with fuzzy matches at top, sorted by score, highlighted.")

if __name__ == "__main__":
    file_sets = [
        {"file1": "./data/internal_data_1.csv",
         "file2": "./data/external_data_1.csv",
         "output_file": "Matched_IndividualData_001.xlsx"}
    ]
    overall_threshold = 69
    start_time = time.time()
    
    for i, file_set in enumerate(file_sets, 1):
        file1 = file_set["file1"]
        file2 = file_set["file2"]
        output_file = file_set["output_file"]
        
        print(f"Processing set {i} of {len(file_sets)}: {file1} and {file2}...")
        try:
            matches = compare_spreadsheets(file1, file2, overall_threshold)
            if matches is not None and not matches.empty:
                print(f"Found {len(matches)} matching rows.")
                print(matches.head())
                export_to_excel(matches, output_file, 'Individual: ID Number')
                print(f"Results exported to {output_file}")
            else:
                print(f"No matching rows found for {file1} and {file2}.")
        except Exception as e:
            print(f"Error processing {file1} and {file2}: {str(e)}")
        
        elapsed_time = time.time() - start_time
        avg_time_per_set = elapsed_time / i
        estimated_time_remaining = avg_time_per_set * (len(file_sets) - i)
        print(f"Estimated time remaining: {timedelta(seconds=int(estimated_time_remaining))}")
        print("-----------------------------------")
    
    total_time = timedelta(seconds=int(time.time() - start_time))
    print(f"Total execution time: {total_time}")
